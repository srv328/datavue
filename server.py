"""Flask сервер для системы управления данными."""
from flask import Flask, request, jsonify, session, Response
from flask_cors import CORS
from functools import wraps
import sqlite3
from datetime import datetime
import csv
import io
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from database import DatabaseManager

app = Flask(__name__)

app.secret_key = 'your-secret-key-here'
CORS(app, supports_credentials=True)


db = DatabaseManager()



def format_datetime(dt_string):
    """Форматирование даты и времени для отображения пользователю."""
    if not dt_string:
        return ""

    try:
        
        dt = datetime.strptime(dt_string, '%Y-%m-%d %H:%M:%S')
        
        return dt.isoformat()
    except (ValueError, TypeError):
        return dt_string



def login_required(f):
    """Декоратор для проверки аутентификации."""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({
            'error': 'Требуется аутентификация'
        }), 401
        return f(*args, **kwargs)
    return decorated_function



def admin_required(f):
    """Декоратор для проверки прав администратора."""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({
            'error': 'Требуется аутентификация'
        }), 401
        if not db.is_admin(session['user_id']):
            return jsonify({
            'error': 'Требуются права администратора'
        }), 403
        return f(*args, **kwargs)
    return decorated_function



@app.route('/api/auth/login', methods=['POST'])


def login():
    """Аутентификация пользователя."""
    data = request.get_json()
    username = data.get('username')
    password = data.get('password')

    if not username or not password:
        return jsonify({
            'error': 'Необходимы имя пользователя и пароль'
        }), 400

    user = db.authenticate_user(username, password)
    if user:
        session['user_id'] = user['id']
        session['username'] = user['username']
        session['role'] = user['role']
        return jsonify({
            'message': 'Успешная аутентификация',
            'user': {
                'id': user['id'],
                'username': user['username'],
                'role': user['role']
            }
        })
    else:
        return jsonify({
            'error': 'Неверные учетные данные'
        }), 401



@app.route('/api/auth/logout', methods=['POST'])


def logout():
    """Выход из системы."""
    session.clear()
    return jsonify({
            'message': 'Успешный выход'
        })



@app.route('/api/auth/me', methods=['GET'])
@login_required


def get_current_user():
    """Получение информации о текущем пользователе."""
    return jsonify({
        'id': session['user_id'],
        'username': session['username'],
        'role': session['role']
    })




@app.route('/api/data-types', methods=['GET'])
@login_required


def get_data_types():
    """Получение списка доступных типов данных."""
    data_types = db.get_data_types(session['user_id'])
    return jsonify(data_types)



@app.route('/api/data-types', methods=['POST'])
@admin_required


def create_data_type():
    """Создание нового типа данных."""
    data = request.get_json()
    name = data.get('name')
    description = data.get('description', '')

    if not name:
        return jsonify({'error': 'Название типа данных обязательно'}), 400

    try:
        data_type_id = db.create_data_type(name, description, session['user_id'])
        return jsonify({
            'message': 'Тип данных создан',
            'data_type_id': data_type_id
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 400



@app.route('/api/data-types/<int:data_type_id>/fields', methods=['GET'])
@login_required


def get_data_fields(data_type_id):
    """Получение полей типа данных."""
    
    if not db.has_permission(session['user_id'], data_type_id, 'read'):
        return jsonify({'error': 'Недостаточно прав для просмотра полей'}), 403

    fields = db.get_data_fields(data_type_id)
    return jsonify(fields)



@app.route('/api/data-types/<int:data_type_id>/has-fields', methods=['GET'])
@login_required


def check_data_type_has_fields(data_type_id):
    """Проверка наличия полей у типа данных."""
    
    if not db.has_permission(session['user_id'], data_type_id, 'read'):
        return jsonify({'error': 'Недостаточно прав для просмотра полей'}), 403

    has_fields = db.has_data_fields(data_type_id)
    return jsonify({'has_fields': has_fields})



@app.route('/api/data-types/<int:data_type_id>/fields', methods=['POST'])
@admin_required


def add_data_field(data_type_id):
    """Добавление поля к типу данных."""
    data = request.get_json()
    field_name = data.get('field_name')
    field_type = data.get('field_type')
    is_required = data.get('is_required', False)
    description = data.get('description', '')
    validation_rules = data.get('validation_rules', '')

    if not field_name or not field_type:
        return jsonify({'error': 'Название и тип поля обязательны'}), 400

    if field_type not in ['text', 'integer', 'decimal', 'date', 'boolean', 'enum', 'coordinates']:
        return jsonify({'error': 'Недопустимый тип поля'}), 400

    try:
        db.add_field_to_data_type(
            data_type_id, field_name, field_type,
            is_required, description, validation_rules
        )
        
        
        if field_type == 'enum':
            enum_values = data.get('enum_values', [])
            if not enum_values or len(enum_values) == 0:
                return jsonify({'error': 'Для типа enum необходимо указать хотя бы одно значение'}), 400
            
            
            fields = db.get_data_fields(data_type_id)
            enum_field = None
            for field in fields:
                if field['field_name'] == field_name and field['field_type'] == 'enum':
                    enum_field = field
                    break
            
            if enum_field:
                db.add_enum_values(enum_field['id'], enum_values)
        
        return jsonify({
            'message': 'Поле добавлено'
        })
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except sqlite3.OperationalError as e:
        if "database is locked" in str(e):
            return jsonify({'error': 'База данных временно заблокирована. Попробуйте еще раз через несколько секунд.'}), 503
        else:
            return jsonify({'error': f'Ошибка базы данных: {str(e)}'}), 500
    except Exception as e:
        return jsonify({'error': f'Неожиданная ошибка: {str(e)}'}), 500



@app.route('/api/data-types/<int:data_type_id>/fields/<int:field_id>/enum-values', methods=['GET'])
@login_required


def get_enum_values(data_type_id, field_id):
    """Получение значений enum поля"""
    try:
        
        fields = db.get_data_fields(data_type_id)
        field = next((f for f in fields if f['id'] == field_id), None)
        
        if not field:
            return jsonify({'error': 'Поле не найдено'}), 404
        
        if field['field_type'] != 'enum':
            return jsonify({'error': 'Поле не является enum типом'}), 400
        
        values = db.get_enum_values(field_id)
        return jsonify({'values': values})
    except Exception as e:
        return jsonify({'error': str(e)}), 500



@app.route('/api/data-types/<int:data_type_id>/fields/<int:field_id>/enum-values', methods=['POST'])
@admin_required


def update_enum_values(data_type_id, field_id):
    """Обновление значений enum поля"""
    try:
        
        fields = db.get_data_fields(data_type_id)
        field = next((f for f in fields if f['id'] == field_id), None)
        
        if not field:
            return jsonify({'error': 'Поле не найдено'}), 404
        
        if field['field_type'] != 'enum':
            return jsonify({'error': 'Поле не является enum типом'}), 400
        
        data = request.get_json()
        enum_values = data.get('values', [])
        
        if not enum_values or len(enum_values) == 0:
            return jsonify({'error': 'Необходимо указать хотя бы одно значение'}), 400
        
        db.add_enum_values(field_id, enum_values)
        return jsonify({'message': 'Значения enum обновлены'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500



@app.route('/api/data-types/<int:data_type_id>/fields/<int:field_id>', methods=['DELETE'])
@admin_required


def delete_data_field(data_type_id, field_id):
    """Удаление поля из типа данных."""
    try:
        db.delete_field_from_data_type(data_type_id, field_id)
        return jsonify({
            'message': 'Поле удалено'
        })
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except sqlite3.OperationalError as e:
        if "database is locked" in str(e):
            return jsonify({'error': 'База данных временно заблокирована. Попробуйте еще раз через несколько секунд.'}), 503
        else:
            return jsonify({'error': f'Ошибка базы данных: {str(e)}'}), 500
    except Exception as e:
        return jsonify({'error': f'Неожиданная ошибка: {str(e)}'}), 500



@app.route('/api/data-types/<int:data_type_id>', methods=['DELETE'])
@admin_required


def delete_data_type(data_type_id):
    """Удаление типа данных."""
    
    if not db.has_permission(session['user_id'], data_type_id, 'admin'):
        return jsonify({'error': 'Недостаточно прав для удаления типа данных'}), 403

    try:
        db.delete_data_type(data_type_id)
        return jsonify({
            'message': 'Тип данных удален'
        })
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except sqlite3.OperationalError as e:
        if "database is locked" in str(e):
            return jsonify({'error': 'База данных временно заблокирована. Попробуйте еще раз через несколько секунд.'}), 503
        else:
            return jsonify({'error': f'Ошибка базы данных: {str(e)}'}), 500
    except Exception as e:
        return jsonify({'error': f'Неожиданная ошибка: {str(e)}'}), 500




@app.route('/api/data/<int:data_type_id>', methods=['GET'])
@login_required


def get_data(data_type_id):
    """Получение записей данных."""
    
    if not db.has_permission(session['user_id'], data_type_id, 'read'):
        return jsonify({
            'error': 'Недостаточно прав для просмотра данных'
        }), 403

    
    limit = request.args.get('limit', 0, type=int)
    offset = request.args.get('offset', 0, type=int)

    records = db.get_data_records(data_type_id, limit, offset)
    return jsonify(records)



@app.route('/api/data/<int:data_type_id>', methods=['POST'])
@login_required


def insert_data(data_type_id):
    """Добавление записи данных."""
    
    if not db.has_permission(session['user_id'], data_type_id, 'write'):
        return jsonify({'error': 'Недостаточно прав для добавления данных'}), 403

    data = request.get_json()

    try:
        
        fields = db.get_data_fields(data_type_id)
        for field in fields:
            if field['field_type'] == 'coordinates':
                field_name = field['field_name']
                if field_name in data:
                    coords = data[field_name]
                    
                    if isinstance(coords, dict):
                        lat = coords.get('latitude')
                        lng = coords.get('longitude')
                        if lat is not None and (lat < -90 or lat > 90):
                            return jsonify({'error': f'Широта должна быть от -90 до 90 градусов (получено: {lat})'}), 400
                        if lng is not None and (lng < -180 or lng > 180):
                            return jsonify({'error': f'Долгота должна быть от -180 до 180 градусов (получено: {lng})'}), 400
                        
                        data[field_name] = json.dumps(coords) if (lat is not None or lng is not None) else None
                    elif isinstance(coords, str):
                        
                        try:
                            coords_dict = json.loads(coords)
                            lat = coords_dict.get('latitude')
                            lng = coords_dict.get('longitude')
                            if lat is not None and (lat < -90 or lat > 90):
                                return jsonify({
                                    'error': 'Широта должна быть от -90 до 90 градусов'
                                }), 400
                            if lng is not None and (lng < -180 or lng > 180):
                                return jsonify({
                                    'error': 'Долгота должна быть от -180 до 180 градусов'
                                }), 400
                        except (json.JSONDecodeError, ValueError):
                            return jsonify({'error': 'Неверный формат координат'}), 400
        
        record_id = db.insert_data_record(data_type_id, data, session['user_id'])
        return jsonify({
            'message': 'Запись добавлена',
            'record_id': record_id
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 400



@app.route('/api/data/<int:data_type_id>/<int:record_id>', methods=['GET'])
@login_required


def get_data_record(data_type_id, record_id):
    """Получение одной записи данных."""
    
    if not db.has_permission(session['user_id'], data_type_id, 'read'):
        return jsonify({'error': 'Недостаточно прав для просмотра данных'}), 403

    try:
        record = db.get_data_record(data_type_id, record_id)
        if not record:
            return jsonify({'error': 'Запись не найдена'}), 404
        return jsonify(record)
    except Exception as e:
        return jsonify({'error': str(e)}), 400



@app.route('/api/data/<int:data_type_id>/<int:record_id>', methods=['PUT'])
@admin_required


def update_data_record(data_type_id, record_id):
    """Обновление записи данных (только для администраторов)."""
    
    if not db.has_permission(session['user_id'], data_type_id, 'admin'):
        return jsonify({'error': 'Недостаточно прав для редактирования данных'}), 403

    data = request.get_json()

    try:
        
        fields = db.get_data_fields(data_type_id)
        for field in fields:
            if field['field_type'] == 'coordinates':
                field_name = field['field_name']
                if field_name in data:
                    coords = data[field_name]
                    
                    if isinstance(coords, dict):
                        lat = coords.get('latitude')
                        lng = coords.get('longitude')
                        if lat is not None and (lat < -90 or lat > 90):
                            return jsonify({'error': f'Широта должна быть от -90 до 90 градусов (получено: {lat})'}), 400
                        if lng is not None and (lng < -180 or lng > 180):
                            return jsonify({'error': f'Долгота должна быть от -180 до 180 градусов (получено: {lng})'}), 400
                        
                        data[field_name] = json.dumps(coords) if (lat is not None or lng is not None) else None
                    elif isinstance(coords, str):
                        
                        try:
                            coords_dict = json.loads(coords)
                            lat = coords_dict.get('latitude')
                            lng = coords_dict.get('longitude')
                            if lat is not None and (lat < -90 or lat > 90):
                                return jsonify({
                                    'error': 'Широта должна быть от -90 до 90 градусов'
                                }), 400
                            if lng is not None and (lng < -180 or lng > 180):
                                return jsonify({
                                    'error': 'Долгота должна быть от -180 до 180 градусов'
                                }), 400
                        except (json.JSONDecodeError, ValueError):
                            return jsonify({'error': 'Неверный формат координат'}), 400
        
        db.update_data_record(data_type_id, record_id, data)
        return jsonify({
            'message': 'Запись обновлена'
        })
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 500



@app.route('/api/data/<int:data_type_id>/<int:record_id>', methods=['DELETE'])
@admin_required


def delete_data_record(data_type_id, record_id):
    """Удаление записи данных (только для администраторов)."""
    
    if not db.has_permission(session['user_id'], data_type_id, 'admin'):
        return jsonify({'error': 'Недостаточно прав для удаления данных'}), 403

    try:
        db.delete_data_record(data_type_id, record_id)
        return jsonify({
            'message': 'Запись удалена'
        })
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 500



@app.route('/api/data/<int:data_type_id>/statistics', methods=['GET'])
@login_required


def get_data_statistics(data_type_id):
    """Получение статистики по данным."""
    
    if not db.has_permission(session['user_id'], data_type_id, 'read'):
        return jsonify({'error': 'Недостаточно прав для просмотра статистики'}), 403

    stats = db.get_data_statistics(data_type_id)
    return jsonify(stats)




@app.route('/api/users', methods=['GET'])
@admin_required


def get_users():
    """Получение списка пользователей."""
    with sqlite3.connect(db.db_path) as conn:
        cursor = conn.cursor()
        cursor.execute('''
            SELECT id, username, full_name, role, created_at, is_active
            FROM users
            ORDER BY created_at DESC
        ''')
        results = cursor.fetchall()
        users = [{
            'id': row[0],
            'username': row[1],
            'full_name': row[2],
            'role': row[3],
            'created_at': format_datetime(row[4]),
            'is_active': row[5]
        } for row in results]
        return jsonify(users)



@app.route('/api/users', methods=['POST'])
@admin_required


def create_user():
    """Создание нового пользователя."""
    data = request.get_json()
    username = data.get('username')
    password = data.get('password')
    role = data.get('role', 'user')
    full_name = data.get('full_name', '')

    if not username or not password:
        return jsonify({'error': 'Имя пользователя и пароль обязательны'}), 400

    if role not in ['admin', 'user']:
        return jsonify({'error': 'Недопустимая роль'}), 400

    try:
        with sqlite3.connect(db.db_path) as conn:
            cursor = conn.cursor()
            password_hash = db.hash_password(password)
            current_time = db.get_current_timestamp()
            cursor.execute('''
                INSERT INTO users (username, password_hash, role, full_name, created_at)
                VALUES (?, ?, ?, ?, ?)
            ''', (username, password_hash, role, full_name, current_time))
            conn.commit()
            return jsonify({
            'message': 'Пользователь создан'
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 400



@app.route('/api/users/generate', methods=['POST'])
@admin_required


def generate_users():
    """Автоматическая генерация пользователей."""
    data = request.get_json()
    count = data.get('count', 1)
    role = data.get('role', 'user')

    if not isinstance(count, int) or count < 1 or count > 100:
        return jsonify({'error': 'Количество должно быть от 1 до 100'}), 400

    if role not in ['admin', 'user']:
        return jsonify({'error': 'Недопустимая роль'}), 400

    try:
        generated_users = db.generate_users(count, role)
        return jsonify({
            'message': f'Создано {len(generated_users)} пользователей',
            'users': generated_users
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 400



@app.route('/api/users/<int:user_id>', methods=['PUT'])
@admin_required


def update_user(user_id):
    """Обновление данных пользователя."""
    data = request.get_json()
    full_name = data.get('full_name', '')

    try:
        with sqlite3.connect(db.db_path) as conn:
            cursor = conn.cursor()
            cursor.execute('''
                UPDATE users
                SET full_name = ?
                WHERE id = ?
            ''', (full_name, user_id))

            if cursor.rowcount == 0:
                return jsonify({'error': 'Пользователь не найден'}), 404

            conn.commit()
            return jsonify({
            'message': 'Пользователь обновлен'
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 400



@app.route('/api/users/<int:user_id>', methods=['DELETE'])
@admin_required


def delete_user(user_id):
    """Удаление пользователя."""
    try:
        with sqlite3.connect(db.db_path) as conn:
            cursor = conn.cursor()

            
            cursor.execute("SELECT username FROM users WHERE id = ?", (user_id,))
            user = cursor.fetchone()
            if not user:
                return jsonify({'error': 'Пользователь не найден'}), 404

            
            if user_id == session['user_id']:
                return jsonify({'error': 'Нельзя удалить собственный аккаунт'}), 400

            
            cursor.execute("DELETE FROM users WHERE id = ?", (user_id,))
            conn.commit()

            return jsonify({
            'message': 'Пользователь удален'
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 400



@app.route('/api/users/<int:user_id>/reset-password', methods=['POST'])
@admin_required


def reset_user_password(user_id):
    """Сброс пароля пользователя."""
    import random
    import string

    try:
        
        new_password = ''.join(random.choices(
            string.ascii_letters + string.digits, k=8
        ))

        with sqlite3.connect(db.db_path) as conn:
            cursor = conn.cursor()
            password_hash = db.hash_password(new_password)
            cursor.execute('''
                UPDATE users
                SET password_hash = ?
                WHERE id = ?
            ''', (password_hash, user_id))
            conn.commit()

            
            cursor.execute('SELECT username FROM users WHERE id = ?', (user_id,))
            result = cursor.fetchone()
            username = result[0] if result else 'unknown'

            return jsonify({
                'message': 'Пароль сброшен',
                'username': username,
                'password': new_password
            })
    except Exception as e:
        return jsonify({'error': str(e)}), 400




@app.route('/api/users/<int:user_id>/permissions', methods=['GET'])
@admin_required


def get_user_permissions(user_id):
    """Получение прав доступа пользователя."""
    try:
        permissions = db.get_user_permissions(user_id)
        return jsonify(permissions)
    except Exception as e:
        return jsonify({'error': str(e)}), 400



@app.route('/api/users/<int:user_id>/permissions/<int:data_type_id>', methods=['POST'])
@admin_required


def grant_permission(user_id, data_type_id):
    """Выдача разрешения пользователю на редактирование типа данных."""
    data = request.get_json()
    permission_type = data.get('permission_type', 'write')

    if permission_type != 'write':
        return jsonify({'error': 'Можно выдавать только право на редактирование'}), 400

    try:
        db.grant_permission(user_id, data_type_id, permission_type, session['user_id'])
        return jsonify({
            'message': 'Разрешение выдано'
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 400



@app.route('/api/users/<int:user_id>/permissions/<int:data_type_id>', methods=['DELETE'])
@admin_required


def revoke_permission(user_id, data_type_id):
    """Отзыв разрешения пользователя."""
    try:
        db.revoke_permission(user_id, data_type_id)
        return jsonify({
            'message': 'Разрешение отозвано'
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 400



@app.route('/api/data-types/<int:data_type_id>/export-csv', methods=['POST'])
@login_required


def export_data_to_csv(data_type_id):
    """Экспорт данных в CSV формат."""
    
    if not db.has_permission(session['user_id'], data_type_id, 'read'):
        return jsonify({'error': 'Недостаточно прав для экспорта данных'}), 403

    try:
        data = request.get_json()
        
        limit = data.get('limit', 100)
        
        include_headers = data.get('include_headers', True)
        
        include_descriptions = data.get('include_descriptions', False)
        
        selected_fields = data.get('selected_fields', None)

        
        data_type = db.get_data_type(data_type_id)
        if not data_type:
            return jsonify({'error': 'Тип данных не найден'}), 404

        
        all_fields = db.get_data_fields(data_type_id)
        if not all_fields:
            return jsonify({'error': 'У типа данных нет полей'}), 400

        
        if selected_fields and len(selected_fields) > 0:
            
            selected_fields_set = set(selected_fields)
            fields = [f for f in all_fields if f['field_name'] in selected_fields_set]
            
            field_order = {name: idx for idx, name in enumerate(selected_fields)}
            fields.sort(key=lambda f: field_order.get(f['field_name'], 999))
        else:
            fields = all_fields

        
        records = db.get_data_records(data_type_id, limit=limit, offset=0)

        
        output = io.StringIO()
        writer = csv.writer(output)

        if include_headers:
            if include_descriptions:
                
                headers = []
                for field in fields:
                    headers.append(f"{field['field_name']} ({field['description']})")
                writer.writerow(headers)
            else:
                
                headers = [field['field_name'] for field in fields]
                writer.writerow(headers)

        
        for record in records:
            row = []
            for field in fields:
                value = record.get(field['field_name'], '')
                
                if value is None:
                    row.append('')
                elif field['field_type'] == 'boolean':
                    row.append('Да' if value else 'Нет')
                elif field['field_type'] == 'coordinates':
                    
                    if isinstance(value, dict):
                        lat = value.get('latitude', '—')
                        lng = value.get('longitude', '—')
                        row.append(f"({lat}, {lng})")
                    elif isinstance(value, str):
                        try:
                            import json
                            coords = json.loads(value)
                            lat = coords.get('latitude', '—')
                            lng = coords.get('longitude', '—')
                            row.append(f"({lat}, {lng})")
                        except Exception:
                            row.append(str(value))
                    else:
                        row.append(str(value))
                elif field['field_type'] in ['integer', 'decimal']:
                    row.append(str(value))
                else:
                    row.append(str(value))
            writer.writerow(row)

        
        output.seek(0)
        csv_data = output.getvalue()
        output.close()

        
        csv_data = '\ufeff' + csv_data

        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        safe_name = data_type['name'].replace(' ', '_').replace('(', '').replace(')', '').replace(',', '').replace(' ', '_')
        filename = f"{safe_name}_{timestamp}.csv"

        return Response(
            csv_data,
            mimetype='text/csv',
            headers={
                'Content-Disposition': f'attachment; filename="{filename}"',
                'Content-Type': 'text/csv; charset=utf-8'
            }
        )

    except Exception as e:
        return jsonify({'error': f'Ошибка экспорта: {str(e)}'}), 500



@app.route('/api/data-types/<int:data_type_id>/export-excel', methods=['POST'])
@login_required


def export_data_to_excel(data_type_id):
    """Экспорт данных в Excel формат."""
    if not db.has_permission(session['user_id'], data_type_id, 'read'):
        return jsonify({'error': 'Недостаточно прав для экспорта данных'}), 403

    try:
        data = request.get_json()
        limit = data.get('limit', 100)
        include_headers = data.get('include_headers', True)
        include_descriptions = data.get('include_descriptions', False)
        
        selected_fields = data.get('selected_fields', None)

        
        data_type = db.get_data_type(data_type_id)
        if not data_type:
            return jsonify({'error': 'Тип данных не найден'}), 404

        
        all_fields = db.get_data_fields(data_type_id)
        if not all_fields:
            return jsonify({'error': 'У типа данных нет полей'}), 400

        
        if selected_fields and len(selected_fields) > 0:
            
            selected_fields_set = set(selected_fields)
            fields = [f for f in all_fields if f['field_name'] in selected_fields_set]
            
            field_order = {name: idx for idx, name in enumerate(selected_fields)}
            fields.sort(key=lambda f: field_order.get(f['field_name'], 999))
        else:
            fields = all_fields

        
        
        actual_limit = limit if limit > 0 else 10000
        records = db.get_data_records(data_type_id, limit=actual_limit, offset=0)

        
        wb = Workbook()
        ws = wb.active
        ws.title = "Данные"

        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        
        if include_headers:
            if include_descriptions:
                headers = []
                for field in fields:
                    headers.append(f"{field['field_name']} ({field['description']})")
            else:
                headers = [field['field_name'] for field in fields]

            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

        for row_idx, record in enumerate(records, 2):
            for col_idx, field in enumerate(fields, 1):
                value = record.get(field['field_name'], '')

                if value is None:
                    cell_value = ''
                elif field['field_type'] == 'boolean':
                    cell_value = 'Да' if value else 'Нет'
                elif field['field_type'] == 'coordinates':
                    if isinstance(value, dict):
                        lat = value.get('latitude', '—')
                        lng = value.get('longitude', '—')
                        cell_value = f"({lat}, {lng})"
                    elif isinstance(value, str):
                        try:
                            import json
                            coords = json.loads(value)
                            lat = coords.get('latitude', '—')
                            lng = coords.get('longitude', '—')
                            cell_value = f"({lat}, {lng})"
                        except Exception:
                            cell_value = str(value)
                    else:
                        cell_value = str(value)
                elif field['field_type'] in ['integer', 'decimal']:
                    cell_value = value
                else:
                    cell_value = str(value)

                ws.cell(row=row_idx, column=col_idx, value=cell_value)

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        excel_data = output.getvalue()
        output.close()

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        safe_name = data_type['name'].replace(' ', '_').replace('(', '').replace(')', '').replace(',', '').replace(' ', '_')
        filename = f"{safe_name}_{timestamp}.xlsx"

        return Response(
            excel_data,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f'attachment; filename="{filename}"',
                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            }
        )

    except Exception as e:
        return jsonify({'error': f'Ошибка экспорта Excel: {str(e)}'}), 500



@app.route('/api/admin/setup', methods=['POST'])


def setup_admin():
    """Настройка учетных данных администратора при первом запуске."""
    try:
        data = request.get_json()
        username = data.get('username', '').strip()
        password = data.get('password', '').strip()
        full_name = data.get('full_name', '').strip()

        if not username or not password:
            return jsonify({'error': 'Логин и пароль обязательны'}), 400

        if len(username) < 3:
            return jsonify({'error': 'Логин должен содержать минимум 3 символа'}), 400

        if len(password) < 6:
            return jsonify({'error': 'Пароль должен содержать минимум 6 символов'}), 400

        with db.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT COUNT(*) FROM users WHERE role = "admin"')
            admin_count = cursor.fetchone()[0]

            if admin_count > 0:
                return jsonify({'error': 'Администратор уже настроен'}), 400

            import hashlib
            password_hash = hashlib.sha256(password.encode()).hexdigest()

            cursor.execute('''
                INSERT INTO users (username, password_hash, full_name, role, created_at)
                VALUES (?, ?, ?, ?, ?)
            ''', (username, password_hash, full_name or 'Администратор системы', 'admin', db.get_current_timestamp()))

            conn.commit()

            return jsonify({
            'message': 'Администратор успешно создан'
        })

    except Exception as e:
        return jsonify({'error': f'Ошибка создания администратора: {str(e)}'}), 500



@app.route('/api/admin/check-setup', methods=['GET'])


def check_admin_setup():
    """Проверка, настроен ли администратор."""
    try:
        with db.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT COUNT(*) FROM users WHERE role = "admin"')
            admin_count = cursor.fetchone()[0]

            return jsonify({'is_setup': admin_count > 0})

    except Exception as e:
        return jsonify({'error': f'Ошибка проверки: {str(e)}'}), 500

if __name__ == '__main__':
    print("Запуск сервера...")
    
    try:
        with db.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT COUNT(*) FROM users WHERE role = "admin"')
            admin_count = cursor.fetchone()[0]
            
        if admin_count == 0:
            print("⚠️  Администратор не настроен. Настройте через веб-интерфейс")
        else:
            print("✅ Администратор настроен")
    except Exception as e:
        print(f"⚠️  Ошибка проверки администратора: {e}")
    
    app.run(debug=True, host='0.0.0.0', port=5000)
